# slab(Bubble) 분기보고 현황 → 엑셀.
#   세로: 조합 포트폴리오 법인, 가로: 제출 여부 / 첨부파일 여부 / 분기보고 기입 내용 전체.
# 실행: python scripts/export_quarterly_status.py [조합명] [연도] [분기]
#       기본값 = SKF4 2026 2분기
# BUBBLE_API_BASE / BUBBLE_API_TOKEN 은 .env.local 에서 읽는다.
import json
import re
import sys
import urllib.request
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")  # Windows 콘솔(cp949)에서 한글/기호 출력

ROOT = Path(__file__).resolve().parent.parent
FUND_NAME = sys.argv[1] if len(sys.argv) > 1 else "SKF4"
YEAR = int(sys.argv[2]) if len(sys.argv) > 2 else 2026
QUARTER = sys.argv[3] if len(sys.argv) > 3 else "2분기"
KST = timezone(timedelta(hours=9))


def load_env():
    env = {}
    for line in (ROOT / ".env.local").read_text(encoding="utf-8").splitlines():
        m = re.match(r"^\s*([A-Z0-9_]+)\s*=\s*(.*)\s*$", line)
        if m:
            env[m.group(1)] = m.group(2).strip("\"'")
    return env


ENV = load_env()
BASE = ENV.get("BUBBLE_API_BASE", "https://slab.sparkerp.co.kr").rstrip("/")
TOKEN = ENV.get("BUBBLE_API_TOKEN", "")


def fetch_all(type_name):
    rows, cursor = [], 0
    while True:
        url = f"{BASE}/api/1.1/obj/{type_name}?cursor={cursor}&limit=100"
        req = urllib.request.Request(url)
        if TOKEN:
            req.add_header("Authorization", f"Bearer {TOKEN}")
        with urllib.request.urlopen(req) as res:
            body = json.loads(res.read().decode("utf-8")).get("response", {})
        batch = body.get("results", [])
        rows += batch
        cursor += len(batch)
        if not batch or body.get("remaining", 0) <= 0:
            return rows


# ---- 값 변환 ---------------------------------------------------------------

def ox(v):
    """리스트/값의 유무를 O/X 로."""
    if isinstance(v, list):
        return "O" if v else "X"
    return "O" if v not in (None, "", False) else "X"


def yn(v):
    return "Y" if v is True else ("N" if v is False else "")


def kdate(v):
    if not v:
        return ""
    try:
        return datetime.fromisoformat(str(v).replace("Z", "+00:00")).astimezone(KST).strftime("%Y-%m-%d")
    except ValueError:
        return str(v)


def txt(v):
    """줄바꿈·중복 공백을 정리한 텍스트(엑셀 셀에서 읽히게)."""
    if v is None:
        return ""
    s = re.sub(r"[ \t]+", " ", str(v))
    return re.sub(r"\n{2,}", "\n", s).strip()


def num(v):
    return v if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def names(ids, lookup):
    if not isinstance(ids, list):
        return ""
    return ", ".join(lookup.get(str(i), str(i)) for i in ids)


def joinlist(v):
    return ", ".join(str(x) for x in v) if isinstance(v, list) else txt(v)


# ---- 수집 -------------------------------------------------------------------

print(f"● slab 조회 중 … ({FUND_NAME} / {YEAR} {QUARTER})")
funds = fetch_all("fund")
companies = fetch_all("company")
quarters = fetch_all("quarterlyupdate")
countries = fetch_all("country")
country_name = {str(c["_id"]): c.get("name", "") for c in countries}

fund = next((f for f in funds if f.get("fund name") == FUND_NAME), None)
if fund is None:
    sys.exit(f"✗ 조합 '{FUND_NAME}' 없음. 가능한 값: {', '.join(f.get('fund name', '') for f in funds)}")

members = [
    c for c in companies
    if str(fund["_id"]) in [str(x) for x in c.get("fund type", []) or []]
    and c.get("company name") != "테스트"
]
report_by_company = {
    str(r.get("company")): r
    for r in quarters
    if r.get("year") == YEAR and r.get("quarter") == QUARTER and r.get("company")
}
print(f"  포트폴리오 {len(members)}개 / 해당 분기 보고행 {sum(1 for c in members if c['_id'] in report_by_company)}개")

# ---- 열 정의 (라벨, 값 추출, 엑셀 표시형식) ---------------------------------
# c = company 레코드, r = 해당 분기 quarterlyupdate 레코드(없으면 {})
WON = "#,##0"
COLS = [
    ("법인명", lambda c, r: txt(c.get("company name")), None),
    ("영문명", lambda c, r: txt(c.get("company name eng")), None),
    ("투자상태", lambda c, r: txt(c.get("company investment status")), None),
    ("분기보고 제출", lambda c, r: "O" if r.get("report made") else "X", None),
    ("제출일", lambda c, r: kdate(r.get("report made")), None),
    ("작성자", lambda c, r: txt(r.get("writer")), None),
    ("메일 발송 횟수", lambda c, r: num(r.get("email sent count")), None),
    # 첨부파일
    ("주주명부", lambda c, r: ox(r.get("shareholders list")), None),
    ("재무제표", lambda c, r: ox(r.get("financial report")), None),
    ("등기부등본", lambda c, r: ox(r.get("company register")), None),
    ("최근 IR 자료", lambda c, r: ox(r.get("updated IR deck")), None),
    # 재무 — 금액 단위는 회사마다 다르다(KRW 274 / USD 67 / SGD·EUR·VND·GBP 소수).
    # 분기보고에는 통화 필드가 없어 company 의 주식 통화를 기준으로 병기한다.
    ("통화(주식 기준)", lambda c, r: txt(c.get("share currency")), None),
    ("보유 현금", lambda c, r: num(r.get("current cash")), WON),
    ("월평균 매출", lambda c, r: num(r.get("monthly average run rate")), WON),
    ("월평균 소모", lambda c, r: num(r.get("monthly average burn rate")), WON),
    ("런웨이(개월)", lambda c, r: num(r.get("runway(month)")), None),
    ("정부지원금 수령", lambda c, r: yn(r.get("government subsidies ")), None),
    ("잔여 정부지원금", lambda c, r: num(r.get("remaining government subsidies")), WON),
    ("정부지원금 종료일", lambda c, r: kdate(r.get("end date of government subsidies")), None),
    # 주식·투자
    ("최신 주당가격", lambda c, r: num(r.get("latest share price")), WON),
    ("최신 발행주식총수", lambda c, r: num(r.get("latest issued share outstanding")), WON),
    ("기업가치(단가×주식수)",
     lambda c, r: (num(r.get("latest share price")) or 0) * (num(r.get("latest issued share outstanding")) or 0) or None,
     WON),
    ("신규 투자유치", lambda c, r: txt(r.get("any new funding round?")), None),
    ("투자유치 진행중", lambda c, r: yn(r.get("funding ongoing")), None),
    ("라운드 시리즈", lambda c, r: txt(r.get("funding series")), None),
    ("조달 금액", lambda c, r: num(r.get("total amount raised")), WON),
    ("라운드 종료일", lambda c, r: kdate(r.get("funding end date")), None),
    ("라운드 주당가격", lambda c, r: num(r.get("share price")), WON),
    ("라운드 발행주식수", lambda c, r: num(r.get("issued share outstanding")), WON),
    ("예상 라운드", lambda c, r: txt(r.get("Expected Funding Round")), None),
    ("예상 라운드 상세", lambda c, r: txt(r.get("Expected Funding Round Details")), None),
    ("주요 주주", lambda c, r: txt(r.get("fund holder details")), None),
    # 조직·해외
    ("인원 수", lambda c, r: num(r.get("head count")), None),
    ("회사 상태", lambda c, r: txt(r.get("company status")), None),
    ("행정구역", lambda c, r: joinlist(r.get("Administrative divisions")), None),
    ("해외 매출", lambda c, r: yn(r.get("overseas revenue")), None),
    ("해외 법인·지사 운영", lambda c, r: yn(r.get("overseas operation")), None),
    ("국내 법인·지점 유무", lambda c, r: yn(r.get("Existence of a domestic corporation or branch")), None),
    ("해외 소재지", lambda c, r: names(r.get("overseas location"), country_name), None),
    ("매출 발생 국가", lambda c, r: names(r.get("revenue source"), country_name), None),
    # 서술형
    ("사업 하이라이트 - 개요", lambda c, r: txt(r.get("business highlight_overview")), None),
    ("사업 하이라이트 - 영업 현황", lambda c, r: txt(r.get("business highlight_business performance")), None),
    ("사업 하이라이트 - 개발 현황", lambda c, r: txt(r.get("business highlight_current_development_progress")), None),
    ("스파크랩 문의사항", lambda c, r: txt(r.get("inquiries to SparkLabs")), None),
    ("최종 수정일", lambda c, r: kdate(r.get("Modified Date")), None),
]

# 미제출을 위로, 그 다음 법인명 가나다순
rows = sorted(
    ((c, report_by_company.get(c["_id"], {})) for c in members),
    key=lambda p: (bool(p[1].get("report made")), p[0].get("company name") or ""),
)

# ---- 엑셀 작성 --------------------------------------------------------------

wb = Workbook()
ws = wb.active
ws.title = f"{YEAR} {QUARTER}"

HEAD_FILL = PatternFill("solid", fgColor="1F3864")
X_FILL = PatternFill("solid", fgColor="FCE4E4")
O_FILL = PatternFill("solid", fgColor="E7F3E7")
OX_COLS = {"분기보고 제출", "주주명부", "재무제표", "등기부등본", "최근 IR 자료"}

ws.append([label for label, _, _ in COLS])
for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = HEAD_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for company, report in rows:
    ws.append([getter(company, report) for _, getter, _ in COLS])

for idx, (label, _, fmt) in enumerate(COLS, start=1):
    letter = get_column_letter(idx)
    if label == "법인명":
        width = 22
    elif label.startswith("사업 하이라이트") or label in ("예상 라운드 상세", "스파크랩 문의사항"):
        width = 60
    else:
        width = max(10, min(20, len(label) * 2 + 2))
    ws.column_dimensions[letter].width = width
    for cell in ws[letter][1:]:
        if fmt:
            cell.number_format = fmt
        if label in OX_COLS:
            cell.alignment = Alignment(horizontal="center")
            cell.fill = X_FILL if cell.value == "X" else O_FILL
        elif width == 60:
            cell.alignment = Alignment(vertical="top")

ws.freeze_panes = "B2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{ws.max_row}"

out = ROOT / "tmp" / f"{FUND_NAME}_{YEAR}_{QUARTER}_분기보고현황.xlsx"
out.parent.mkdir(exist_ok=True)
wb.save(out)

submitted = sum(1 for _, r in rows if r.get("report made"))
print(f"✓ 저장: {out}")
print(f"  법인 {len(rows)}개 · 제출 {submitted} / 미제출 {len(rows) - submitted}")
